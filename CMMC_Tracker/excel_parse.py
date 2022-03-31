from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from apps.home.models import Controls

book = load_workbook('All_CMMC_v1_Controls.xlsx')
ws = book.active

for row in range(1, 174):
    for col in range(1, 5):
        char = get_column_letter(col)

        print(ws[char + str(row)].value)